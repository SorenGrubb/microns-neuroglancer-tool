"""
v2 -- supersedes build_new_cells_outcome_report.py's output. Soren looked at the outcome and
realized the 363 "new cells" were actually merged-segment/multi-nucleus blobs; the addition was
reverted (revert_new_cells_join.py). Cross-checking against the existing "mergeddata" sub-cell
block explains why: 321 of 363 already exist there in some form (156 exact-duplicate override,
23 new-sub-cell-at-a-known-merged-index, 165 standalone-matches-a-sub-cell-elsewhere), leaving 42
that don't match anything already recorded. Soren's decision (2026-08-07 follow-up): park the 42
for now, don't add anything -- this report just documents the corrected picture for the record.

Run: python3 build_new_cells_outcome_report_v2.py
Output: all_own_data_new_cells_outcome.xlsx (in this same folder) -- OVERWRITES the v1 outcome file
already in the project folder (same filename), since that one incorrectly said the 363 were added.
"""
import re, json
import openpyxl
from openpyxl.styles import Font, PatternFill

F = "/sessions/optimistic-intelligent-carson/mnt/MICrONS cell classifier/microns-neuroglancer-tool/ujump.html"
html = open(F, encoding="utf-8").read()

def get_block(name):
    m = re.search(r'<script type="application/json" id="%s">(.*?)</script>' % re.escape(name), html, re.S)
    return json.loads(m.group(1))

MD = get_block("mergeddata")
merged_idx_set = set(int(k) for k in MD.keys())
all_subs = [(int(k), s) for k, subs in MD.items() for s in subs]

def cent_match(a, b, tol=2):
    if a is None or b is None:
        return False
    return all(abs(a[i] - b[i]) <= tol for i in range(3))

def vox_str(p):
    if not p:
        return ""
    return "%d, %d, %d" % (round(p[0]), round(p[1]), round(p[2]))

plan = json.load(open("/tmp/dash/new_cells_join_plan.json"))
add = plan["add"]
excluded = plan["excluded_red"]
override_entries = [e for e in add if e["kind"] == "override"]
standalone_entries = [e for e in add if e["kind"] == "standalone"]

rows = []
for e in override_entries:
    idx = e["nucdata_index"]
    cent = [round(c) for c in e["centriole"]] if e["centriole"] else None
    if idx not in merged_idx_set:
        status = "NOT in mergeddata -- residual, parked"
    else:
        exact = any(s["type"] == e["type"] and cent_match(cent, s.get("centriole")) for s in MD[str(idx)])
        status = "Exact duplicate of an existing mergeddata sub-cell" if exact else "New sub-cell type at an already-known merged index -- residual, parked"
    rows.append([e["row"], e["type"], "override", status, vox_str(e["nuc_voxel"]), vox_str(e["centriole"])])

for e in standalone_entries:
    cent = [round(c) for c in e["centriole"]] if e["centriole"] else None
    found = any(s["type"] == e["type"] and cent_match(cent, s.get("centriole")) for _, s in all_subs)
    status = "Matches an existing mergeddata sub-cell record elsewhere in the dataset" if found else "No match anywhere -- residual, parked"
    rows.append([e["row"], e["type"], "standalone", status, vox_str(e["nuc_voxel"]), vox_str(e["centriole"])])

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Outcome (NOT added)"
ws1.append(["xlsx row", "cell type", "original join kind", "status vs. existing mergeddata", "nucleus coordinate (voxel)", "centriole coordinate (voxel)"])
for c in ws1[1]:
    c.font = Font(bold=True)
residual_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
for r in sorted(rows, key=lambda r: r[0]):
    ws1.append(r)
    if "residual" in r[3]:
        for cell in ws1[ws1.max_row]:
            cell.fill = residual_fill
widths1 = [10, 18, 14, 46, 26, 26]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[chr(64 + i)].width = w

ws2 = wb.create_sheet("Excluded (red-flagged)")
ws2.append(["xlsx row", "cell type", "nucleus coordinate (voxel x,y,z)", "reason"])
for c in ws2[1]:
    c.font = Font(bold=True)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
for e in sorted(excluded, key=lambda e: e["row"]):
    ws2.append([e["row"], e["type"], vox_str(e["nuc_voxel"]), "Red text/fill in All own data.xlsx -- not trusted, not added"])
    for cell in ws2[ws2.max_row]:
        cell.fill = red_fill
widths2 = [10, 20, 26, 50]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[chr(64 + i)].width = w

ws3 = wb.create_sheet("Summary")
ws3.append(["Metric", "Count"])
for c in ws3[1]:
    c.font = Font(bold=True)
n_dup = sum(1 for r in rows if "duplicate" in r[3] or "elsewhere" in r[3])
n_residual = sum(1 for r in rows if "residual" in r[3])
summary_rows = [
    ("Candidates re-derived from All own data.xlsx", len(add) + len(excluded)),
    ("Excluded -- red text/fill (not trusted)", len(excluded)),
    ("NOTHING added to ujump.html this pass (addition was reverted)", ""),
    ("  ...already present as a mergeddata sub-cell (duplicate info, correctly not new)", n_dup),
    ("  ...residual -- not matched anywhere, parked per Soren's decision (2026-08-07)", n_residual),
]
for r in summary_rows:
    ws3.append(list(r))
ws3.column_dimensions["A"].width = 70
ws3.column_dimensions["B"].width = 12

wb.save("/tmp/dash/all_own_data_new_cells_outcome.xlsx")
print("Saved. Duplicates:", n_dup, "Residual (parked):", n_residual, "Excluded:", len(excluded))
