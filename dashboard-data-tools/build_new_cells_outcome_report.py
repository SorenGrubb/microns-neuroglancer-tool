"""
Replaces all_own_data_new_cells_to_review.xlsx (which framed the 462 candidates as "pending review")
with a report of what actually happened, per Soren's follow-up: "The 462 candidates are solid, add
them as new verified cells, I trust them. However, those of them that were red text or red
background in the all own data, I don't trust and should not be included."

363 were added to ujump.html's owndata/cildata (180 as overrides onto a previously-unclaimed
MICrONS-detected nucleus, 183 as standalone points); 99 were excluded for red flags and NOT added.
This is now a record of the outcome, not a to-do list.

Run: python3 build_new_cells_outcome_report.py
Output: all_own_data_new_cells_outcome.xlsx (in this same folder) -- Soren should replace the old
all_own_data_new_cells_to_review.xlsx in the project folder with this one.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill

plan = json.load(open("/tmp/dash/new_cells_join_plan.json"))
add = plan["add"]
excluded = plan["excluded_red"]

def vox_str(p):
    if not p:
        return ""
    return "%d, %d, %d" % (round(p[0]), round(p[1]), round(p[2]))

wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "Added as new cells"
ws1.append(["xlsx row", "cell type", "outcome", "MICrONS nucleus ID (if override)",
            "nucleus coordinate (voxel x,y,z)", "has centriole", "has cilium", "authors/citation"])
for c in ws1[1]:
    c.font = Font(bold=True)
for e in sorted(add, key=lambda e: e["row"]):
    ws1.append([
        e["row"], e["type"], e["kind"],
        e.get("nucleus_id", ""),
        vox_str(e["nuc_voxel"]),
        "Yes" if e["centriole"] else "",
        "Yes" if (e["cil_start"] and e["cil_end"]) else "",
        e["authors"],
    ])
widths1 = [10, 20, 12, 22, 26, 12, 12, 60]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[chr(64 + i)].width = w

ws2 = wb.create_sheet("Excluded (red-flagged)")
ws2.append(["xlsx row", "cell type", "nucleus coordinate (voxel x,y,z)", "reason"])
for c in ws2[1]:
    c.font = Font(bold=True)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
for e in sorted(excluded, key=lambda e: e["row"]):
    ws2.append([e["row"], e["type"], vox_str(e["nuc_voxel"]), "Red text/fill in All own data.xlsx -- not trusted, not added"])
    for cell in ws2[ws2.max_row]:
        cell.fill = red_fill
widths2 = [10, 20, 26, 50]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[chr(64 + i)].width = w

ws3 = wb.create_sheet("Summary")
ws3.append(["Metric", "Count"])
for c in ws3[1]:
    c.font = Font(bold=True)
n_override = sum(1 for e in add if e["kind"] == "override")
n_standalone = sum(1 for e in add if e["kind"] == "standalone")
n_cent = sum(1 for e in add if e["centriole"])
n_cil = sum(1 for e in add if e["cil_start"] and e["cil_end"])
rows = [
    ("Total candidates re-derived from All own data.xlsx", len(add) + len(excluded)),
    ("Excluded -- red text/fill (not trusted)", len(excluded)),
    ("Added as new verified cells", len(add)),
    ("  ...as OVERRIDE (landed on a free MICrONS-detected nucleus)", n_override),
    ("  ...as STANDALONE (no matching MICrONS nucleus nearby)", n_standalone),
    ("  ...with a centriole coordinate", n_cent),
    ("  ...with full cilium (base+tip) coordinates", n_cil),
]
for r in rows:
    ws3.append(list(r))
ws3.column_dimensions["A"].width = 60
ws3.column_dimensions["B"].width = 12

wb.save("/tmp/dash/all_own_data_new_cells_outcome.xlsx")
print("Saved /tmp/dash/all_own_data_new_cells_outcome.xlsx")
print("Added:", len(add), "(override:", n_override, "standalone:", n_standalone, ") -- Excluded:", len(excluded))
